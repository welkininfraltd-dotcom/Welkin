"""Bank Statement PDF → Excel Translator.

Parses Indian Bank statement PDFs and produces an Excel file matching
the '525 RJ.xls' template format used by Welkin Builders.

Usage:
    python bank_statement_parser.py <input_pdf> [output_xlsx]

If output_xlsx is not specified, it defaults to the PDF filename with .xlsx extension.
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ── Data Models ────────────────────────────────────────────────────

@dataclass
class Transaction:
    value_date: str = ""
    post_date: str = ""
    remitter_branch: str = ""
    description: str = ""
    cheque_ref: str = ""
    debit: Optional[float] = None
    credit: Optional[float] = None
    balance: str = ""


@dataclass
class StatementMetadata:
    bank_name: str = "INDIAN BANK"
    branch: str = ""
    ifsc: str = ""
    address: str = ""
    branch_code: str = ""
    account_number: str = ""
    product_type: str = ""
    account_holder: str = ""
    holder_address: list[str] = field(default_factory=list)
    email: str = ""
    statement_date: str = ""
    cleared_balance: str = ""
    uncleared_amount: str = ""
    drawing_power: str = ""
    interest_rate: str = ""
    period_from: str = ""
    period_to: str = ""


# ── PDF Parsing ────────────────────────────────────────────────────

def parse_metadata_from_text(full_text: str) -> StatementMetadata:
    """Extract header/metadata from the first page text."""
    meta = StatementMetadata()

    # Period
    m = re.search(r"from\s+(\d{2}/\d{2}/\d{4})\s+to\s+(\d{2}/\d{2}/\d{4})", full_text)
    if m:
        meta.period_from = m.group(1)
        meta.period_to = m.group(2)

    # Account number
    m = re.search(r"Account Number\s*:\s*(\d+)", full_text)
    if m:
        meta.account_number = m.group(1)

    # IFSC
    m = re.search(r"IFSC CODE\s*:\s*(\w+)", full_text)
    if m:
        meta.ifsc = m.group(1)

    # Branch
    lines = full_text.split("\n")
    for i, line in enumerate(lines):
        if "INDIAN BANK" in line:
            if i + 1 < len(lines):
                meta.branch = lines[i + 1].strip()
            break

    # Branch code
    m = re.search(r"Branch Code\s*:\s*(\d+)", full_text)
    if m:
        meta.branch_code = m.group(1)

    # Product type
    m = re.search(r"Product type\s*:\s*(.+)", full_text)
    if m:
        meta.product_type = m.group(1).strip()

    # Address line
    m = re.search(r"G-1.*Pradesh", full_text)
    if m:
        meta.address = m.group(0).strip()

    # Account holder
    m = re.search(r"WELKIN BUILDERS.*?LTD", full_text)
    if m:
        meta.account_holder = m.group(0).strip()

    # Email
    m = re.search(r"Email\s*:\s*(\S+)", full_text)
    if m:
        meta.email = m.group(1)

    # Statement date
    m = re.search(r"Statement Date\s*:(.*)", full_text)
    if m:
        meta.statement_date = m.group(1).strip()

    # Balances
    m = re.search(r"Cleared Balance\s*:\s*([\d.]+)", full_text)
    if m:
        meta.cleared_balance = m.group(1)

    m = re.search(r"Uncleared Amount\s*:\s*([\d.]+)", full_text)
    if m:
        meta.uncleared_amount = m.group(1)

    m = re.search(r"Drawing Power\s*:\s*([\d.]+)", full_text)
    if m:
        meta.drawing_power = m.group(1)

    m = re.search(r"Interest Rate\s*:\s*([\d.]+)", full_text)
    if m:
        meta.interest_rate = m.group(1)

    return meta


def _clean_date(raw: str) -> str:
    """Convert '01/04\\n/2026' or '01/04/2026' to '01/04/2026'."""
    if not raw:
        return ""
    cleaned = raw.replace("\n", "").replace(" ", "").strip()
    # Handle format like '01/04/2026'
    m = re.match(r"(\d{2}/\d{2})/(\d{4})", cleaned)
    if m:
        return f"  {m.group(1)}/{m.group(2)}"
    return f"  {cleaned}"


def _clean_balance(raw: str) -> str:
    """Convert '6852654.36\\nCR' to '6852654.36CR'."""
    if not raw:
        return ""
    return raw.replace("\n", "").replace(" ", "").strip()


def _parse_amount(raw: str) -> Optional[float]:
    """Parse amount string to float, return None if empty."""
    if not raw or not raw.strip():
        return None
    cleaned = raw.replace(",", "").replace("\n", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _extract_meaningful_description(raw_desc: str) -> tuple[str, str]:
    """Extract a meaningful short description and reference from raw bank description.
    
    Returns (description, cheque_ref) matching the XLS template style.
    The description is the transaction type/purpose, cheque_ref is the party name.
    
    Indian Bank PDF splits names across lines, so we join lines before extracting.
    """
    if not raw_desc:
        return ("", "")

    # Join all lines — names are split across newlines in the PDF
    # e.g., "B\nHARAT S" → "BHARAT S"
    raw_joined = raw_desc.replace("\n", "")
    # Also keep a space-joined version for readability
    lines = [l.strip() for l in raw_desc.split("\n") if l.strip()]
    full = " ".join(lines)

    party_name = ""
    desc = ""

    # ── BALANCE B/F ──
    if "BALANCE B/F" in full:
        return ("BALANCE B/F", "")

    # ── COMMISSION CHARGES (bank fees) ──
    if "COMMISSION CHARGES" in full or ("CHARGES" in full and "IMPS" in full):
        return ("Bank Charges", "IMPS Commission")

    # ── NEFT transactions ──
    # In the joined text: NEFT/BANKCODE/REF/NAME/. Txn Amt.
    # e.g., "NEFT/JSFB/IDIBN52026040136079975/BHARAT S/. Txn Amt."
    neft_match = re.search(r"NEFT/\w+/[A-Z0-9]+/([^/]+?)/?\.\s*Txn", raw_joined)
    if neft_match:
        party_name = neft_match.group(1).strip()
        # Clean trailing dots/spaces
        party_name = party_name.rstrip(". ")
        desc = "NEFT Transfer"
        return (desc, party_name)

    # Also try space-joined version for NEFT
    neft_match2 = re.search(r"NEFT/\w+/[A-Z0-9 ]+/([^/]+?)/?\.\s*Txn", full)
    if neft_match2:
        party_name = neft_match2.group(1).strip().rstrip(". ")
        desc = "NEFT Transfer"
        return (desc, party_name)

    # ── IMPS transactions ──
    # Joined: /IMPS/P2A/ref/account/BANK/Name TRANSFER TO account
    # e.g., "UBIN/Arju ArjunSingh TRANSFER"
    imps_match = re.search(r"/IMPS/P2A/\d+/\d+/(\w+)/(.+?)(?:TRANSFER|$)", raw_joined)
    if imps_match:
        party_name = imps_match.group(2).strip()
        # Insert spaces before capitals for camelCase names: "ArjunSingh" → "Arjun Singh"
        party_name = re.sub(r"([a-z])([A-Z])", r"\1 \2", party_name)
        # Remove trailing "im" prefix duplicates like "Nitin im Nitin Choudhary"
        dup_match = re.match(r"(\w+)\s+im\s+(.+)", party_name, re.IGNORECASE)
        if dup_match:
            party_name = dup_match.group(2).strip()
        desc = "IMPS Transfer"
        return (desc, party_name)

    # Try space-joined IMPS
    imps_match2 = re.search(r"/IMPS/P2A/[\d ]+/[\d ]+/(\w+)/(.+?)(?:\s*TRANSFER|$)", full)
    if imps_match2:
        party_name = imps_match2.group(2).strip()
        party_name = re.sub(r"([a-z])([A-Z])", r"\1 \2", party_name)
        dup_match = re.match(r"(\w+)\s+im\s+(.+)", party_name, re.IGNORECASE)
        if dup_match:
            party_name = dup_match.group(2).strip()
        desc = "IMPS Transfer"
        return (desc, party_name)

    # ── Online Transfer with name embedded ──
    # Pattern: "TRANSFER Transfer 73515Onkar Chouhan OnkarChauhan TRANSFER TO..."
    online_match = re.search(r"TRANSFER\s+(?:Transfer\s*)?(\d+)([A-Za-z][A-Za-z\s]+?)(?:TRANSFER TO|$)", full)
    if online_match:
        party_name = online_match.group(2).strip()
        # Remove duplicated camelCase version
        party_name = re.sub(r"([a-z])([A-Z])", r"\1 \2", party_name)
        # Take first occurrence of the name (before any duplicate)
        parts = party_name.split()
        if len(parts) >= 2:
            party_name = " ".join(parts[:2])
        desc = "Online Transfer"
        return (desc, party_name)

    # ── Online Transfer: TRANSFER TO account Name ──
    online_match2 = re.search(r"TRANSFER TO\s+\d+\s+([A-Za-z][A-Za-z\s]+)", full)
    if online_match2:
        party_name = online_match2.group(1).strip()
        party_name = re.sub(r"\s+TRANSFER.*", "", party_name).strip()
        desc = "Online Transfer"
        return (desc, party_name)

    # ── BY TRANSFER (credit/deposit) ──
    if "BY TRANSFER" in full or "BY CLG" in full:
        by_match = re.search(r"(?:BY TRANSFER|BY CLG)\s+(\d+)", full)
        ref = by_match.group(1) if by_match else ""
        # Try to find a name after the reference
        name_match = re.search(r"(?:BY TRANSFER|BY CLG)\s+\d*\s*([A-Za-z][A-Za-z\s&]+)", full)
        if name_match:
            ref = name_match.group(1).strip()
        desc = "Credit Transfer"
        return (desc, ref)

    # ── UPI transactions ──
    upi_match = re.search(r"UPI/.*?/([A-Za-z\s]+?)(?:/|@|\s*$)", full)
    if upi_match:
        party_name = upi_match.group(1).strip()
        desc = "UPI Transfer"
        return (desc, party_name)

    # ── Deposit/Credit ──
    if "DEPOSIT" in full or "BY CLG" in full:
        dep_match = re.search(r"(?:DEPOSIT|BY CLG).*?([A-Z][A-Za-z\s&]+)", full)
        if dep_match:
            party_name = dep_match.group(1).strip()
        desc = "Deposit"
        return (desc, party_name)

    # ── Generic WITHDRAWAL with TRANSFER TO ──
    if "WITHDRAWAL" in full:
        # Try joined version for name extraction
        wdl_name = re.search(r"TRANSFER\s+TO\s+\d+\s+([A-Za-z][A-Za-z\s]+)", full)
        if wdl_name:
            party_name = wdl_name.group(1).strip()
            party_name = re.sub(r"\s+TRANSFER.*", "", party_name).strip()
        # If still no name, try the raw joined for RTGS/other patterns
        if not party_name or party_name == "TO":
            wdl_name2 = re.search(r"(?:RTGS|NEFT|IMPS)/.*?/([A-Za-z][A-Za-z\s&]+?)(?:/|\.\s*Txn|TRANSFER|$)", raw_joined)
            if wdl_name2:
                party_name = wdl_name2.group(1).strip()
        # Last resort: look for CHQ or cheque pattern
        if not party_name or party_name == "TO":
            chq_match = re.search(r"CHQ\s+(\w+)", full)
            if chq_match:
                party_name = chq_match.group(1)
                desc = "Cheque Transfer"
                return (desc, party_name)
            party_name = ""
        desc = "Withdrawal"
        return (desc, party_name)

    # ── Fallback ──
    desc = lines[0] if lines else ""
    return (desc, party_name)


def _get_remitter_branch(raw: str) -> str:
    """Clean up remitter branch text."""
    if not raw:
        return ""
    cleaned = raw.replace("\n", " ").strip()
    return f"  {cleaned}"


def extract_transactions_from_pdf(pdf_path: str) -> tuple[StatementMetadata, list[Transaction]]:
    """Parse the entire PDF and return metadata + list of transactions."""
    transactions: list[Transaction] = []
    metadata: Optional[StatementMetadata] = None

    with pdfplumber.open(pdf_path) as pdf:
        # Get metadata from first page
        first_page_text = pdf.pages[0].extract_text() or ""
        metadata = parse_metadata_from_text(first_page_text)

        # Extract tables from all pages
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 8:
                        continue

                    # Skip header rows
                    if row[0] and ("Value" in row[0] or "Date" in row[0]):
                        continue

                    value_date = _clean_date(row[0] or "")
                    post_date = _clean_date(row[1] or "")
                    remitter_branch = _get_remitter_branch(row[2] or "")
                    raw_description = row[3] or ""
                    cheque_no = row[4] or ""
                    debit = _parse_amount(row[5] or "")
                    credit = _parse_amount(row[6] or "")
                    balance = _clean_balance(row[7] or "")

                    # Parse description into meaningful text
                    desc_text, ref_text = _extract_meaningful_description(raw_description)

                    # For BALANCE B/F row
                    if "BALANCE B/F" in raw_description:
                        txn = Transaction(
                            value_date="  ",
                            post_date="  ",
                            remitter_branch="  ",
                            description=" BALANCE B/F",
                            cheque_ref=" ",
                            debit=None,
                            credit=None,
                            balance=balance,
                        )
                        transactions.append(txn)
                        continue

                    txn = Transaction(
                        value_date=value_date,
                        post_date=post_date,
                        remitter_branch=remitter_branch,
                        description=desc_text,
                        cheque_ref=ref_text or cheque_no.replace("\n", " ").strip(),
                        debit=debit,
                        credit=credit,
                        balance=balance,
                    )
                    transactions.append(txn)

    return metadata, transactions


# ── Excel Output ───────────────────────────────────────────────────

def write_excel(
    metadata: StatementMetadata,
    transactions: list[Transaction],
    output_path: str,
) -> None:
    """Write the parsed data to an Excel file matching the 525 RJ.xls template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement of Account"

    # Styles
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    data_font = Font(size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Column widths
    col_widths = [14, 14, 28, 20, 25, 15, 15, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Header section (rows 1-20) ──
    row = 1
    # Row 1: empty
    row += 1

    # Row 2-8: Bank info (centered in column D)
    ws.cell(row=2, column=4, value=f"{metadata.bank_name}  ").font = title_font
    ws.cell(row=3, column=4, value=metadata.branch).font = data_font
    ws.cell(row=4, column=4, value=f"IFSC CODE :{metadata.ifsc}").font = data_font
    ws.cell(row=5, column=4, value=metadata.address).font = data_font
    ws.cell(row=6, column=4, value=f"Branch Code :{metadata.branch_code}").font = data_font
    ws.cell(row=7, column=4, value=f"Account Number : {metadata.account_number}").font = data_font
    ws.cell(row=8, column=4, value=f"Product type :  {metadata.product_type}").font = data_font

    # Row 9-13: Account holder info (column A)
    ws.cell(row=9, column=1, value=metadata.account_holder).font = header_font
    ws.cell(row=10, column=1, value="705 PUKHRAJ  CORPORATE").font = data_font
    ws.cell(row=11, column=1, value="IN FRONT OF  NAVLAKHA BUS STAND").font = data_font
    ws.cell(row=12, column=1, value="INDORE").font = data_font
    ws.cell(row=13, column=1, value="MADHYA PRADESH - 452001").font = data_font

    # Row 14-18: Additional info
    ws.cell(row=14, column=1, value=f"Email : {metadata.email}").font = data_font
    ws.cell(row=15, column=1, value=f"Statement Date :{metadata.statement_date}").font = data_font
    ws.cell(row=16, column=1, value=f"Cleared Balance :{metadata.cleared_balance}").font = data_font
    ws.cell(row=17, column=1, value=f"Uncleared Amount :{metadata.uncleared_amount}").font = data_font
    ws.cell(row=18, column=1, value=f"Drawing Power :{metadata.drawing_power}").font = data_font
    ws.cell(row=19, column=1, value=f"Interest Rate : {metadata.interest_rate}").font = data_font

    # Row 20: Statement period
    ws.cell(row=20, column=1, value=f"Statement of Account from {metadata.period_from} to {metadata.period_to}").font = header_font

    # Row 21: empty
    # Row 22: Column headers
    headers = ["Value Date", "Post Date", "Remitter Branch", "Description",
               "Chq No/REF No/UTR No", "Debit Amount", "Credit Amount", "Balance"]
    header_row = 22
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Transaction rows ──
    total_debit = 0.0
    total_credit = 0.0
    data_row = header_row + 1

    for txn in transactions:
        ws.cell(row=data_row, column=1, value=txn.value_date).font = data_font
        ws.cell(row=data_row, column=2, value=txn.post_date).font = data_font
        ws.cell(row=data_row, column=3, value=txn.remitter_branch).font = data_font
        ws.cell(row=data_row, column=4, value=txn.description).font = data_font
        ws.cell(row=data_row, column=5, value=txn.cheque_ref).font = data_font

        if txn.debit is not None:
            ws.cell(row=data_row, column=6, value=f"{txn.debit:.2f}").font = data_font
            total_debit += txn.debit
        else:
            ws.cell(row=data_row, column=6, value=" ").font = data_font

        if txn.credit is not None:
            ws.cell(row=data_row, column=7, value=f"{txn.credit:.2f}").font = data_font
            total_credit += txn.credit
        else:
            ws.cell(row=data_row, column=7, value=" ").font = data_font

        ws.cell(row=data_row, column=8, value=txn.balance).font = data_font

        # Apply borders
        for col in range(1, 9):
            ws.cell(row=data_row, column=col).border = thin_border

        data_row += 1

    # ── Totals row ──
    ws.cell(row=data_row, column=1, value="Total").font = header_font
    ws.cell(row=data_row, column=6, value=f"{total_debit:,.2f}").font = header_font
    ws.cell(row=data_row, column=7, value=f"{total_credit:,.2f}").font = header_font
    for col in range(1, 9):
        ws.cell(row=data_row, column=col).border = thin_border
    data_row += 1

    # ── Footer ──
    ws.cell(row=data_row + 1, column=1,
            value=f"  * Statement Downloaded By {metadata.account_holder} on {metadata.statement_date} "
                  "Unless a constituent notifies the Bank ").font = Font(size=8, italic=True)
    ws.cell(row=data_row + 2, column=1,
            value="immediately of any discrepancy found by him/her in this statement of a/c, "
                  "it will be taken that he has found the a/c correct.").font = Font(size=8, italic=True)
    ws.cell(row=data_row + 3, column=1,
            value="END OF STATEMENT - from Internet Banking").font = Font(size=8, italic=True)

    wb.save(output_path)
    print(f"✓ Excel saved: {output_path}")
    print(f"  Transactions: {len(transactions)}")
    print(f"  Total Debit:  ₹{total_debit:,.2f}")
    print(f"  Total Credit: ₹{total_credit:,.2f}")


# ── Main ───────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python bank_statement_parser.py <input_pdf> [output_xlsx]")
        print("\nExample:")
        print("  python bank_statement_parser.py StatementOfAccount_6086423525_01052026_143335.pdf")
        sys.exit(1)

    input_pdf = sys.argv[1]
    if not Path(input_pdf).exists():
        print(f"Error: File not found: {input_pdf}")
        sys.exit(1)

    # Default output name
    if len(sys.argv) >= 3:
        output_xlsx = sys.argv[2]
    else:
        output_xlsx = Path(input_pdf).stem + "_parsed.xlsx"

    print(f"Parsing: {input_pdf}")
    metadata, transactions = extract_transactions_from_pdf(input_pdf)
    print(f"Found {len(transactions)} transactions")
    print(f"Period: {metadata.period_from} to {metadata.period_to}")
    print(f"Account: {metadata.account_number}")

    write_excel(metadata, transactions, output_xlsx)


if __name__ == "__main__":
    main()
